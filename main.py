import os
import re
import platform
import subprocess
import time
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from urllib.parse import urlparse

# Input and output Excel files
INPUT_EXCEL = "database.xlsx"
OUTPUT_EXCEL = "output.xlsx"


# Read Excel file
df = pd.read_excel(INPUT_EXCEL)

df.columns = df.columns.str.strip()


if "Link" not in df.columns:
    raise ValueError(f"Excel file must contain 'Link' column. Found: {list(df.columns)}")

if "Quantity" not in df.columns:
    raise ValueError(f"Excel file must contain 'Quantity' column. Found: {list(df.columns)}")


contents = []
status = []
prices = []
total_prices = []

def get_page(url):

    headers = {
        "User-Agent": 
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 "
        "Chrome/120.0 Safari/537.36",

        "Accept":
        "text/html,application/xhtml+xml",

        "Accept-Language":
        "en-US,en;q=0.9"
    }


    for attempt in range(2):

        try:

            response = requests.get(
                url,
                headers=headers,
                timeout=30
            )

            response.raise_for_status()

            return response.text


        except Exception as e:

            print(
                f"Attempt {attempt+1}/2 failed: {url}"
            )


            if attempt < 1:

                time.sleep(3)

            else:

                raise e
            
def find_price(soup, text):

    selectors = [
        ("span", {"class": "price"}, None),
        ("span", {"itemprop": "price"}, None),
        ("meta", {"name": "price"}, "content"),
        ("meta", {"name": "twitter:data1"}, "content"),
        ("meta", {"property": "product:price:amount"}, "content"),
        ("meta", {"property": "og:price:amount"}, "content"),

        ("span", {"class": "old-prices"}, None),
        ("span", {"class": "new-price"}, None),
        ("div", {"class": "current-price"}, None),
        ("div", {"class": "setak-price-amount"}, None),
        ("div", {"id": "setak-price-display"}, None),
        ("p", {"class": "price"}, None),
        ("ins", {}, None),
        ("bdi", {}, None),
        ("meta", {"itemprop": "price"}, "content"),
        ("meta", {"property": "product:price"}, "content"),
  

        ("div", {"class": "price"}, None),
        ("div", {"class": "product-price"}, None),
        ("div", {"class": "woocommerce-Price-amount"}, None),

        ("strong", {"class": "price"}, None),
        ("h2", {}, None),
        ("h3", {}, None),
    ]

    # shop.qom-elec.ir style
    element = soup.find("div", id="setak-price-display")

    if element:

        txt = element.get_text(" ", strip=True)

        m = re.search(r'[\d۰-۹][\d۰-۹,٬.]*', txt)

        if m:

            currency = "ریال"

            # Check SVG icon
            use = element.find("use")

            if use:
                href = (use.get("href") or use.get("xlink:href") or "").lower()

                if "#toman" in href:
                    currency = "تومان"

                elif "#rial" in href:
                    currency = "ریال"

            return m.group(), currency
    # bahar-enclosure style
    element = soup.select_one(
        "span.woocommerce-Price-amount"
    )

    if element:

        txt = element.get_text(" ", strip=True)

        m = re.search(
            r'([\d۰-۹,٬]+)\s*(ریال|تومان)',
            txt
        )

        if m:
            return m.group(1), m.group(2)
    
    #buybestelectronic style
    for element in soup.find_all(
        class_=lambda c: c and "new-price" in c
    ):

        txt = element.get_text(" ", strip=True)

        m = re.search(
            r'([\d۰-۹,٬]+)\s*(ریال|تومان)',
            txt
        )

        if m:
            return m.group(1), m.group(2)

        
    # javanelec style price
    price_box = soup.find(
        lambda tag:
        tag.name in ["div","span"]
        and "ریال" in tag.get_text()
    )

    if price_box:

        txt = price_box.get_text(" ", strip=True)

        m = re.search(
            r'([\d۰-۹,٬]+)\s*(ریال|تومان)',
            txt
        )

        if m:
            return m.group(1), m.group(2)
    
    # Search all matching elements
    for tag, attrs, attribute in selectors:

        for element in soup.find_all(tag, attrs=attrs):

            if attribute:
                price_text = element.get(attribute, "")

                # Also check visible text for currency
                visible_text = " ".join(element.stripped_strings)

                if "تومان" in visible_text:
                    price_text += " تومان"

                elif "ریال" in visible_text:
                    price_text += " ریال"

            else:
                price_text = " ".join(element.stripped_strings)

            if not price_text:
                continue

            price_text = price_text.replace("\xa0", " ")

            m = re.search(
                r'([\d۰-۹][\d۰-۹,٬]*)\D*(ریال|تومان)',
                price_text
            )

            if m:
                return m.group(1), m.group(2)

            # Meta tags that contain only numbers
            m = re.search(r'[\d۰-۹][\d۰-۹,٬]*', price_text)

            if m:
                currency = "تومان" if "تومان" in price_text else "ریال"
                return m.group(), currency

    # Search every element on page
    for element in soup.find_all():

        price_text = " ".join(element.stripped_strings)

        m = re.search(
            r'([\d۰-۹][\d۰-۹,٬]*)\D*(ریال|تومان)',
            price_text
        )

        if m:
            return m.group(1), m.group(2)
    # -------- More HTML patterns --------

    # itemprop="price"
    element = soup.select_one('[itemprop="price"]')

    if element:

        content_price = element.get("content", "")
        visible_price = element.get_text(" ", strip=True)

        price = content_price or visible_price

        m = re.search(
            r'[\d۰-۹][\d۰-۹,٬.]*',
            price
        )

        if m:

            if "تومان" in visible_price:
                currency = "تومان"

            elif "ریال" in visible_price:
                currency = "ریال"

            else:
                currency = "ریال"

            return m.group(), currency

    # Any element with class containing "price"
    for element in soup.select('[class*="price"]'):
        txt = element.get_text(" ", strip=True)
        if txt:
            m = re.search(r'([\d۰-۹][\d۰-۹,٬.]*)\s*(ریال|تومان)?', txt)
            if m:
                currency = m.group(2) if m.group(2) else ("تومان" if "تومان" in txt else "ریال")
                return m.group(1), currency

    # data-price attributes
    for attr in ["data-price-amount","data-price","data-product-price","price","content"]:
        element = soup.find(attrs={attr: True})
        if element:
            value = element.get(attr)
            if value:
                m = re.search(r'[\d۰-۹][\d۰-۹,٬.]*', value)
                if m:

                    parent_text = element.get_text(" ", strip=True)

                    if "تومان" in parent_text:
                        currency = "تومان"

                    elif "ریال" in parent_text:
                        currency = "ریال"

                    else:
                        currency = "ریال"

                    return m.group(), currency

    # JSON-LD Product price
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            import json
            data = json.loads(script.string)

            def search_price(obj):
                if isinstance(obj, dict):
                    if "price" in obj:
                        return str(obj["price"])
                    for v in obj.values():
                        r = search_price(v)
                        if r:
                            return r
                elif isinstance(obj, list):
                    for item in obj:
                        r = search_price(item)
                        if r:
                            return r
                return None

            price = search_price(data)
            if price:
                m = re.search(r'[\d۰-۹][\d۰-۹,٬.]*', price)
                if m:
                    return m.group(), "ریال"

        except:
            pass

    # OpenGraph price
    element = soup.find("meta", property="product:price:amount")
    if element:
        return element.get("content"), "ریال"

    element = soup.find("meta", property="og:price:amount")
    if element:
        return element.get("content"), "ریال"

    # Twitter price
    element = soup.find("meta", attrs={"name": "twitter:data1"})
    if element:
        txt = element.get("content", "")
        m = re.search(r'[\d۰-۹][\d۰-۹,٬.]*', txt)
        if m:
            currency = "تومان" if "تومان" in txt else "ریال"
            return m.group(), currency
    # Final fallback: search whole page text
    m = re.search(
        r'([\d۰-۹][\d۰-۹,٬]*)\D*(ریال|تومان)',
        text
    )

    if m:
        return m.group(1), m.group(2)

    return None

for index, link in enumerate(df["Link"]):

    try:

        if pd.isna(link):
            contents.append("")
            status.append("NO")
            prices.append("")
            total_prices.append("")
            continue


        # Read HTML
        if str(link).startswith(("http://", "https://")):

            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) "
                              "Chrome/120.0 Safari/537.36",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9"
            }

            html = get_page(link)

        else:

            with open(link, "r", encoding="utf-8") as f:
                html = f.read()



        # Parse HTML
        soup = BeautifulSoup(html, "html.parser")


        for tag in soup(["script", "style"]):
            tag.decompose()


        text = soup.get_text(separator=" ", strip=True)


        contents.append(text)


        match = find_price(soup, text)

        if match:

            status.append("OK")

            price = match[0]
            currency = match[1]

            price = price.translate(
                str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")
            )

            price = price.replace(",", "").replace("٬", "")

            price = float(price)

            if currency == "تومان":
                price *= 10

            prices.append(int(price))

            quantity = df.loc[index, "Quantity"]

            if pd.isna(quantity):
                quantity = 1

            total_prices.append(int(price * float(quantity)))

        else:

            status.append("NO")
            prices.append("")
            total_prices.append("")


    except Exception as e:

        print(f"Error processing {link}: {e}")

        contents.append(f"ERROR: {e}")
        status.append("ERROR")
        prices.append("")
        total_prices.append("")



# Add columns
# df["Content"] = contents   # Don't save this column
df["Status"] = status
df["Price"] = prices
df["Total Price"] = total_prices

# Save Excel
df.to_excel(OUTPUT_EXCEL, index=False)



# Open workbook
wb = load_workbook(OUTPUT_EXCEL)
ws = wb.active



# -------------------------------
# Make Link clickable
# -------------------------------

link_col = None

for cell in ws[1]:

    if cell.value == "Link":
        link_col = cell.column
        break


if link_col:

    for row in range(2, ws.max_row + 1):

        cell = ws.cell(row=row, column=link_col)

        if cell.value:

            cell.hyperlink = str(cell.value)
            cell.style = "Hyperlink"



# -------------------------------
# Color Status column only
# -------------------------------

status_col = None

for cell in ws[1]:
    if cell.value == "Status":
        status_col = cell.column
        break

green_fill = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid"
)

red_fill = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)

if status_col:
    for row in range(2, ws.max_row + 1):

        status_cell = ws.cell(row=row, column=status_col)

        if status_cell.value == "OK":
            status_cell.fill = green_fill

        elif status_cell.value in ["NO", "ERROR"]:
            status_cell.fill = red_fill

# -------------------------------
# Color Link column by website
# -------------------------------

from urllib.parse import urlparse

site_colors = {}

colors = [
    "D9D2E9",  # Purple
    "CFE2F3",  # Light Blue
    "FCE5CD",  # Orange
    "FFF2CC",  # Yellow
    "D0E0E3",  # Gray Blue
    "EAD1DC",  # Pink Purple
    "B6D7A8",  # Light Green
    "F4CCCC",  # Soft Pink
    "C9DAF8",  # Sky Blue
    "FFD966",  # Gold
    "B4A7D6",  # Violet
    "A2C4C9"   # Teal
]

if link_col:

    color_index = 0

    for row in range(2, ws.max_row + 1):

        link_cell = ws.cell(row=row, column=link_col)

        if link_cell.value:

            try:
                domain = urlparse(str(link_cell.value)).netloc

                # remove www.
                domain = domain.replace("www.", "")

                if domain not in site_colors:

                    site_colors[domain] = PatternFill(
                        start_color=colors[color_index % len(colors)],
                        end_color=colors[color_index % len(colors)],
                        fill_type="solid"
                    )

                    color_index += 1


                # Color only Link cell
                link_cell.fill = site_colors[domain]

            except:
                pass
# -------------------------------
# Add Grand Total
# -------------------------------

total_col = None

for cell in ws[1]:

    if cell.value == "Total Price":
        total_col = cell.column
        break



if total_col:

    grand_total = 0


    for row in range(2, ws.max_row + 1):

        value = ws.cell(row=row, column=total_col).value


        if value:

            try:
                grand_total += float(value)

            except:
                pass



    last_row = ws.max_row + 1


    ws.cell(row=last_row, column=total_col - 1).value = "Grand Total"

    ws.cell(row=last_row, column=total_col).value = grand_total

# -------------------------------
# Format Price columns with comma
# -------------------------------

price_col = None
total_price_col = None

for cell in ws[1]:

    if cell.value == "Price":
        price_col = cell.column

    if cell.value == "Total Price":
        total_price_col = cell.column


if price_col:

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=price_col).number_format = '#,##0'


if total_price_col:

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=total_price_col).number_format = '#,##0'

# Save final workbook
wb.save(OUTPUT_EXCEL)


print(f"Done! Output saved to {OUTPUT_EXCEL}")



# -------------------------------
# Auto open Excel
# -------------------------------

try:

    if platform.system() == "Windows":

        os.startfile(OUTPUT_EXCEL)


    elif platform.system() == "Darwin":

        subprocess.call(["open", OUTPUT_EXCEL])


    else:

        subprocess.call(["xdg-open", OUTPUT_EXCEL])


except Exception as e:

    print(f"Could not open Excel file automatically: {e}")